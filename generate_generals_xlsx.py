import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ==============================================================================
# КОНФИГУРАЦИЯ И ДАННЫЕ
# ==============================================================================
DEFS_PATH = 'defs.json'
OUTPUT_DIR = 'game_info'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'generals_120_stats.xlsx')
COEFF_120 = 317.52

# Ручной маппинг способностей на основе анализа defs.json
ABILITY_MAP = {
    "bioplasm": {
        "desc": "Биоплазма: Урон врагам + Лечение своих (5 сек)",
        "calc": lambda dmg: f"Урон: {round(dmg*2.5)}x5, Хил: {round(dmg*2.0)}x5",
        "total": lambda dmg: f"Σ Урон: {round(dmg*12.5)}, Σ Хил: {round(dmg*10.0)}"
    },
    "acidCorrosion": {
        "desc": "Кислотная коррозия: Снижает урон врагов на 40% (11 сек)",
        "calc": lambda dmg: "Дебафф: -40% DMG врагам",
        "total": lambda dmg: "-"
    },
    "brainstorm": {
        "desc": "Мозговой штурм: Урон + Бафф урона союзникам (40%) + Стан (2 сек)",
        "calc": lambda dmg: f"Урон: {round(dmg*2.0)}, Бафф: +40% DMG своим",
        "total": lambda dmg: f"Σ Урон: {round(dmg*2.0)}"
    },
    "molotovCoctail": {
        "desc": "Коктейль Молотова: Огненный урон по площади (11 сек)",
        "calc": lambda dmg: f"Урон: {round(dmg*0.2)} за сек (11 тактов)",
        "total": lambda dmg: f"Σ Урон: {round(dmg*2.2)}"
    },
    "fuselOilVapor": {
        "desc": "Сивушные пары: Снижает скорость атаки врагов на 50% (5 сек)",
        "calc": lambda dmg: "Дебафф: -50% Скорость атаки",
        "total": lambda dmg: "-"
    },
    "excitationEnergy": {
        "desc": "Энергия возбуждения: Увеличивает урон всех союзников на 50%",
        "calc": lambda dmg: "Бафф: +50% DMG союзникам",
        "total": lambda dmg: "-"
    },
    "frostStrike": {
        "desc": "Морозный удар: Снижает скорость атаки врагов на 50% (11 сек)",
        "calc": lambda dmg: "Дебафф: -50% Скорость атаки",
        "total": lambda dmg: "-"
    },
    "rocketAttack": {
        "desc": "Ракетная атака: Наносит 300% урона",
        "calc": lambda dmg: f"Урон: {round(dmg*3.0)}",
        "total": lambda dmg: f"Σ Урон: {round(dmg*3.0)}"
    },
    "cumulativeProjectile": {
        "desc": "Кумулятивный снаряд: 450% урона + Стан (1 сек)",
        "calc": lambda dmg: f"Урон: {round(dmg*4.5)}, Стан: 1с",
        "total": lambda dmg: f"Σ Урон: {round(dmg*4.5)}"
    },
    "swarm": {
        "desc": "Рой: Увеличивает шанс крита союзников на 15% (17 сек)",
        "calc": lambda dmg: "Бафф: +15% CRIT союзникам",
        "total": lambda dmg: "-"
    },
    "bloom": {
        "desc": "Цветение: Мощное лечение союзников (15 сек)",
        "calc": lambda dmg: f"Лечение: {round(dmg*0.6)} за сек (15 тактов)",
        "total": lambda dmg: f"Σ Хил: {round(dmg*9.0)}"
    },
    "skyPulse": {
        "desc": "Небесный импульс: 450% урона + Стан (1 сек)",
        "calc": lambda dmg: f"Урон: {round(dmg*4.5)}, Стан: 1с",
        "total": lambda dmg: f"Σ Урон: {round(dmg*4.5)}"
    },
    "sniperShot": {
        "desc": "Снайперский выстрел: 1200% урона по цели с макс. HP",
        "calc": lambda dmg: f"Урон: {round(dmg*12.0)}",
        "total": lambda dmg: f"Σ Урон: {round(dmg*12.0)}"
    }
}

def load_defs():
    with open(DEFS_PATH, 'r', encoding='utf-8') as f:
        f.readline()
        return json.loads(f.read())

def generate():
    data = load_defs()
    generals = data['generals']['generals']
    
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.title = "Generals Comparison 120"
    
    # Заголовки (строки)
    row_headers = [
        "ГЕНЕРАЛ",
        "Атака (Damage)",
        "Здоровье (Health)",
        "Критический удар (Crit)",
        "Уворот (Dodge)",
        "Анти-Крит (AntiCrit)",
        "Анти-Уворот (AntiDodge)",
        "---",
        "СПОСОБНОСТЬ",
        "Суть способности",
        "Откат (сек)",
        "Стоимость (Energy)",
        "Эффективность (за такт)",
        "СУММАРНЫЙ ПОКАЗАТЕЛЬ"
    ]
    
    for r_idx, header in enumerate(row_headers, 1):
        cell = ws.cell(row=r_idx, column=1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')

    col_idx = 2
    # Сортируем и исключаем дубликаты
    sorted_gids = sorted([g for g in generals.keys() if g != 'infectedCat'])
    
    for gid in sorted_gids:
        gdef = generals[gid]
        stats = gdef.get('stats', {})
        
        # Точный расчет статов 120 уровня
        dmg = round(float(stats.get('damage', '0').replace(',', '.')) * COEFF_120)
        hp = round(float(stats.get('health', '0').replace(',', '.')) * COEFF_120)
        crit = round(float(stats.get('crit', '0').replace(',', '.')) * COEFF_120)
        dodge = round(float(stats.get('dodge', '0').replace(',', '.')) * COEFF_120)
        acrit = round(float(stats.get('antiCrit', '0').replace(',', '.')) * COEFF_120)
        adodge = round(float(stats.get('antiDodge', '0').replace(',', '.')) * COEFF_120)
        
        ab_id = gdef.get('ability', '')
        ab_cfg = ABILITY_MAP.get(ab_id, {"desc": ab_id, "calc": lambda x: "-", "total": lambda x: "-"})
        
        # Заполнение столбца
        ws.cell(row=1, column=col_idx, value=gid.upper()).alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        ws.cell(row=2, column=col_idx, value=dmg)
        ws.cell(row=3, column=col_idx, value=hp)
        ws.cell(row=4, column=col_idx, value=crit)
        ws.cell(row=5, column=col_idx, value=dodge)
        ws.cell(row=6, column=col_idx, value=acrit)
        ws.cell(row=7, column=col_idx, value=adodge)
        ws.cell(row=8, column=col_idx, value="")
        
        ws.cell(row=9, column=col_idx, value=ab_id.upper()).font = Font(italic=True)
        ws.cell(row=10, column=col_idx, value=ab_cfg['desc']).alignment = Alignment(wrapText=True)
        ws.cell(row=11, column=col_idx, value=int(data['abilities'].get(ab_id, {}).get('cooldown', '0s').replace('s', '')))
        ws.cell(row=12, column=col_idx, value=int(float(data['abilities'].get(ab_id, {}).get('energyCost', '0').replace(',', '.'))))
        
        ws.cell(row=13, column=col_idx, value=ab_cfg['calc'](dmg))
        ws.cell(row=14, column=col_idx, value=ab_cfg['total'](dmg)).font = Font(bold=True)
        
        col_idx += 1

    last_col = get_column_letter(col_idx - 1)
    
    # Визуальная настройка
    ws.row_dimensions[1].height = 100 # Высота для вертикальных имен
    ws.column_dimensions['A'].width = 22
    for c in range(2, col_idx):
        ws.column_dimensions[get_column_letter(c)].width = 12 # Увеличил ширину для читаемости абилок

    # Градиенты для основных статов
    color_rule = ColorScaleRule(start_type='min', start_color='F8696B',
                                 mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                 end_type='max', end_color='63BE7B')
    
    for r in [2, 3, 4, 5, 6, 7]:
        ws.conditional_formatting.add(f'B{r}:{last_col}{r}', color_rule)

    wb.save(OUTPUT_FILE)
    print(f"Отчет успешно создан в папке {OUTPUT_DIR}: {os.path.basename(OUTPUT_FILE)}")

if __name__ == "__main__":
    generate()
